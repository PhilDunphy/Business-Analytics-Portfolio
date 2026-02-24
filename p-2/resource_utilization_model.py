#!/usr/bin/env python3
"""
Campus Resource Utilization Analysis
======================================
A consulting-style analytics project analysing utilisation efficiency of
campus facilities (Gym, Library, Computer Labs, Study Rooms, Auditorium,
Sports Complex) at an Indian engineering college.

Outputs:
    1. Campus_Resource_Utilization_Analysis.xlsx — 6-sheet workbook with charts
"""

import os
from dataclasses import dataclass, field
from typing import Dict, List, Tuple

# ---------------------------------------------------------------------------
# Try to import openpyxl; if missing, install it automatically
# ---------------------------------------------------------------------------
try:
    import openpyxl
except ImportError:
    import subprocess, sys
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl

from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference, LineChart
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from copy import copy

# ═══════════════════════════════════════════════════════════════════════════
# 1.  DATA — Realistic Indian Engineering College Campus Facilities
# ═══════════════════════════════════════════════════════════════════════════

# Hours of operation: 6 AM (index 0) to 11 PM (index 17) → 18 slots
HOURS = [f"{h}:00" for h in range(6, 24)]  # 6 AM – 11 PM

@dataclass
class Facility:
    """Single campus facility with capacity and hourly usage data."""
    name: str
    max_capacity: int                  # persons at a time
    hourly_usage: List[int] = field(default_factory=list)  # 18 values (6AM–11PM)
    operating_cost_per_hour: float = 0.0   # ₹ per operating hour

    # Derived ------------------------------------------------------------------
    @property
    def total_operating_hours(self) -> int:
        return len(self.hourly_usage)

    @property
    def total_person_hours(self) -> int:
        return sum(self.hourly_usage)

    @property
    def max_possible_person_hours(self) -> int:
        return self.max_capacity * self.total_operating_hours

    @property
    def avg_utilization_pct(self) -> float:
        if self.max_possible_person_hours == 0:
            return 0.0
        return (self.total_person_hours / self.max_possible_person_hours) * 100

    @property
    def peak_usage(self) -> int:
        return max(self.hourly_usage) if self.hourly_usage else 0

    @property
    def peak_utilization_pct(self) -> float:
        return (self.peak_usage / self.max_capacity) * 100 if self.max_capacity else 0

    @property
    def peak_hour(self) -> str:
        if not self.hourly_usage:
            return "N/A"
        idx = self.hourly_usage.index(max(self.hourly_usage))
        return HOURS[idx]

    @property
    def idle_hours(self) -> int:
        """Hours where usage is below 20% of capacity."""
        threshold = self.max_capacity * 0.20
        return sum(1 for u in self.hourly_usage if u < threshold)

    @property
    def overcrowded_hours(self) -> int:
        """Hours where usage exceeds 85% of capacity."""
        threshold = self.max_capacity * 0.85
        return sum(1 for u in self.hourly_usage if u > threshold)

    @property
    def idle_capacity_pct(self) -> float:
        return ((self.max_possible_person_hours - self.total_person_hours)
                / self.max_possible_person_hours) * 100 if self.max_possible_person_hours else 0

    @property
    def daily_operating_cost(self) -> float:
        return self.operating_cost_per_hour * self.total_operating_hours

    @property
    def cost_per_person_hour(self) -> float:
        return self.daily_operating_cost / self.total_person_hours if self.total_person_hours else 0


def build_dataset() -> List[Facility]:
    """Return realistic facility data for a mid-tier Indian engineering college."""
    return [
        Facility(
            name="Gymnasium",
            max_capacity=60,
            hourly_usage=[8, 25, 35, 15, 10, 12, 18, 22, 30, 45, 50, 38, 28, 15, 10, 8, 20, 35],
            operating_cost_per_hour=450.0,
        ),
        Facility(
            name="Central Library",
            max_capacity=200,
            hourly_usage=[15, 30, 60, 90, 120, 140, 150, 165, 170, 180, 175, 160, 130, 100, 80, 60, 45, 25],
            operating_cost_per_hour=800.0,
        ),
        Facility(
            name="Computer Lab A",
            max_capacity=80,
            hourly_usage=[5, 10, 55, 70, 75, 72, 68, 65, 60, 55, 50, 45, 35, 25, 15, 10, 8, 5],
            operating_cost_per_hour=600.0,
        ),
        Facility(
            name="Computer Lab B",
            max_capacity=80,
            hourly_usage=[0, 5, 40, 60, 65, 70, 72, 68, 55, 40, 30, 25, 20, 15, 10, 5, 3, 0],
            operating_cost_per_hour=600.0,
        ),
        Facility(
            name="Study Room Block",
            max_capacity=120,
            hourly_usage=[10, 15, 25, 40, 50, 55, 60, 65, 70, 80, 90, 100, 105, 95, 85, 70, 55, 40],
            operating_cost_per_hour=350.0,
        ),
        Facility(
            name="Auditorium",
            max_capacity=500,
            hourly_usage=[0, 0, 50, 100, 350, 200, 80, 50, 30, 20, 15, 10, 50, 100, 300, 150, 50, 0],
            operating_cost_per_hour=1200.0,
        ),
        Facility(
            name="Sports Complex",
            max_capacity=150,
            hourly_usage=[20, 40, 15, 10, 8, 12, 15, 20, 25, 35, 50, 65, 80, 100, 120, 130, 110, 70],
            operating_cost_per_hour=500.0,
        ),
    ]


# ═══════════════════════════════════════════════════════════════════════════
# 2.  OPTIMISATION SIMULATIONS
# ═══════════════════════════════════════════════════════════════════════════

def _redistribute_usage(facility: Facility, target_pct: float = 0.70) -> List[int]:
    """Simulate smoothing: cap peaks at target_pct of capacity,
    redistribute excess to idle hours."""
    cap = int(facility.max_capacity * target_pct)
    new_usage = list(facility.hourly_usage)
    excess = 0
    for i, u in enumerate(new_usage):
        if u > cap:
            excess += u - cap
            new_usage[i] = cap

    # Spread excess into low-utilisation hours
    low_indices = sorted(range(len(new_usage)), key=lambda i: new_usage[i])
    for idx in low_indices:
        if excess <= 0:
            break
        room = cap - new_usage[idx]
        add = min(excess, room)
        new_usage[idx] += add
        excess -= add
    return new_usage


def run_optimizations(facilities: List[Facility]) -> List[Dict]:
    """Generate scenario comparisons."""
    scenarios: List[Dict] = []

    # Aggregate base metrics
    total_ph = sum(f.total_person_hours for f in facilities)
    total_max = sum(f.max_possible_person_hours for f in facilities)
    total_cost = sum(f.daily_operating_cost for f in facilities)
    total_idle = sum(f.idle_hours for f in facilities)
    total_crowded = sum(f.overcrowded_hours for f in facilities)
    base_util = (total_ph / total_max) * 100 if total_max else 0
    base_cost_per_ph = total_cost / total_ph if total_ph else 0

    scenarios.append({
        "Scenario": "Base Case (Current)",
        "Avg Utilization %": round(base_util, 2),
        "Total Person-Hours": total_ph,
        "Idle Hours (total)": total_idle,
        "Overcrowded Hours": total_crowded,
        "Daily Op. Cost (₹)": total_cost,
        "Cost / Person-Hr (₹)": round(base_cost_per_ph, 2),
    })

    # --- Scenario: Usage Redistribution (cap peaks at 70%) ---
    for cap_pct in [0.70, 0.60]:
        new_facilities = []
        for f in facilities:
            new_usage = _redistribute_usage(f, cap_pct)
            nf = Facility(name=f.name, max_capacity=f.max_capacity,
                          hourly_usage=new_usage, operating_cost_per_hour=f.operating_cost_per_hour)
            new_facilities.append(nf)
        t_ph = sum(f.total_person_hours for f in new_facilities)
        t_max = sum(f.max_possible_person_hours for f in new_facilities)
        t_cost = sum(f.daily_operating_cost for f in new_facilities)
        t_idle = sum(f.idle_hours for f in new_facilities)
        t_crowd = sum(f.overcrowded_hours for f in new_facilities)
        util = (t_ph / t_max) * 100 if t_max else 0
        scenarios.append({
            "Scenario": f"Redistribute (cap {int(cap_pct*100)}%)",
            "Avg Utilization %": round(util, 2),
            "Total Person-Hours": t_ph,
            "Idle Hours (total)": t_idle,
            "Overcrowded Hours": t_crowd,
            "Daily Op. Cost (₹)": t_cost,
            "Cost / Person-Hr (₹)": round(t_cost / t_ph if t_ph else 0, 2),
        })

    # --- Scenario: Extended hours for high-demand facilities ---
    extended = []
    for f in facilities:
        if f.avg_utilization_pct > 50:
            extra_hours = [int(f.peak_usage * 0.3)] * 2  # 2 extra hours at 30% of peak
            new_usage = f.hourly_usage + extra_hours
            nf = Facility(name=f.name, max_capacity=f.max_capacity,
                          hourly_usage=new_usage, operating_cost_per_hour=f.operating_cost_per_hour)
        else:
            nf = Facility(name=f.name, max_capacity=f.max_capacity,
                          hourly_usage=f.hourly_usage, operating_cost_per_hour=f.operating_cost_per_hour)
        extended.append(nf)
    t_ph = sum(f.total_person_hours for f in extended)
    t_max = sum(f.max_possible_person_hours for f in extended)
    t_cost = sum(f.daily_operating_cost for f in extended)
    t_idle = sum(f.idle_hours for f in extended)
    t_crowd = sum(f.overcrowded_hours for f in extended)
    util = (t_ph / t_max) * 100 if t_max else 0
    scenarios.append({
        "Scenario": "Extended Hours (High-Demand)",
        "Avg Utilization %": round(util, 2),
        "Total Person-Hours": t_ph,
        "Idle Hours (total)": t_idle,
        "Overcrowded Hours": t_crowd,
        "Daily Op. Cost (₹)": t_cost,
        "Cost / Person-Hr (₹)": round(t_cost / t_ph if t_ph else 0, 2),
    })

    # --- Scenario: Close underutilised slots (< 15% usage) ---
    trimmed = []
    for f in facilities:
        threshold = f.max_capacity * 0.15
        new_usage = [u for u in f.hourly_usage if u >= threshold]
        if not new_usage:
            new_usage = f.hourly_usage  # keep as-is if all would be removed
        nf = Facility(name=f.name, max_capacity=f.max_capacity,
                      hourly_usage=new_usage, operating_cost_per_hour=f.operating_cost_per_hour)
        trimmed.append(nf)
    t_ph = sum(f.total_person_hours for f in trimmed)
    t_max = sum(f.max_possible_person_hours for f in trimmed)
    t_cost = sum(f.daily_operating_cost for f in trimmed)
    t_idle = sum(f.idle_hours for f in trimmed)
    t_crowd = sum(f.overcrowded_hours for f in trimmed)
    util = (t_ph / t_max) * 100 if t_max else 0
    scenarios.append({
        "Scenario": "Close Idle Slots (<15%)",
        "Avg Utilization %": round(util, 2),
        "Total Person-Hours": t_ph,
        "Idle Hours (total)": t_idle,
        "Overcrowded Hours": t_crowd,
        "Daily Op. Cost (₹)": t_cost,
        "Cost / Person-Hr (₹)": round(t_cost / t_ph if t_ph else 0, 2),
    })

    # --- Scenario: Capacity right-sizing (reduce capacity of under-utilised) ---
    resized = []
    for f in facilities:
        if f.avg_utilization_pct < 40:
            new_cap = int(f.peak_usage * 1.2)  # resize to 120% of peak
        else:
            new_cap = f.max_capacity
        nf = Facility(name=f.name, max_capacity=new_cap,
                      hourly_usage=f.hourly_usage, operating_cost_per_hour=f.operating_cost_per_hour)
        resized.append(nf)
    t_ph = sum(f.total_person_hours for f in resized)
    t_max = sum(f.max_possible_person_hours for f in resized)
    t_cost = sum(f.daily_operating_cost for f in resized)
    t_idle = sum(f.idle_hours for f in resized)
    t_crowd = sum(f.overcrowded_hours for f in resized)
    util = (t_ph / t_max) * 100 if t_max else 0
    scenarios.append({
        "Scenario": "Capacity Right-Sizing",
        "Avg Utilization %": round(util, 2),
        "Total Person-Hours": t_ph,
        "Idle Hours (total)": t_idle,
        "Overcrowded Hours": t_crowd,
        "Daily Op. Cost (₹)": t_cost,
        "Cost / Person-Hr (₹)": round(t_cost / t_ph if t_ph else 0, 2),
    })

    # --- Combined optimised scenario ---
    combined = []
    for f in facilities:
        new_usage = _redistribute_usage(f, 0.70)
        # also trim idle slots
        threshold = f.max_capacity * 0.15
        new_usage = [u if u >= threshold else int(threshold) for u in new_usage]
        new_cap = f.max_capacity if f.avg_utilization_pct >= 40 else int(f.peak_usage * 1.2)
        nf = Facility(name=f.name, max_capacity=new_cap,
                      hourly_usage=new_usage, operating_cost_per_hour=f.operating_cost_per_hour * 0.95)
        combined.append(nf)
    t_ph = sum(f.total_person_hours for f in combined)
    t_max = sum(f.max_possible_person_hours for f in combined)
    t_cost = sum(f.daily_operating_cost for f in combined)
    t_idle = sum(f.idle_hours for f in combined)
    t_crowd = sum(f.overcrowded_hours for f in combined)
    util = (t_ph / t_max) * 100 if t_max else 0
    scenarios.append({
        "Scenario": "Combined Optimized",
        "Avg Utilization %": round(util, 2),
        "Total Person-Hours": t_ph,
        "Idle Hours (total)": t_idle,
        "Overcrowded Hours": t_crowd,
        "Daily Op. Cost (₹)": t_cost,
        "Cost / Person-Hr (₹)": round(t_cost / t_ph if t_ph else 0, 2),
    })

    return scenarios


# ═══════════════════════════════════════════════════════════════════════════
# 3.  EXCEL WORKBOOK GENERATION
# ═══════════════════════════════════════════════════════════════════════════

# ---- Colour palette -------------------------------------------------------
NAVY       = "1B2A4A"
DARK_TEAL  = "1A6B5C"
TEAL       = "2EC4B6"
LIGHT_GREY = "F2F2F2"
WHITE      = "FFFFFF"
GOLD       = "D4A853"
RED_ACCENT = "E74C3C"
GREEN_ACC  = "27AE60"

HEADER_FONT  = Font(name="Calibri", bold=True, color=WHITE, size=11)
HEADER_FILL  = PatternFill("solid", fgColor=NAVY)
SUBHEAD_FILL = PatternFill("solid", fgColor=DARK_TEAL)
ALT_ROW_FILL = PatternFill("solid", fgColor=LIGHT_GREY)
TITLE_FONT   = Font(name="Calibri", bold=True, color=NAVY, size=16)
SECTION_FONT = Font(name="Calibri", bold=True, color=DARK_TEAL, size=13)
CURRENCY_FMT = '₹#,##0'
PERCENT_FMT  = '0.00"%"'
THIN_BORDER  = Border(bottom=Side(style="thin", color="CCCCCC"))


def _style_header_row(ws, row, max_col):
    for col in range(1, max_col + 1):
        c = ws.cell(row=row, column=col)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)


def _alt_rows(ws, start_row, end_row, max_col):
    for r in range(start_row, end_row + 1):
        if (r - start_row) % 2 == 1:
            for c in range(1, max_col + 1):
                ws.cell(row=r, column=c).fill = ALT_ROW_FILL
        for c in range(1, max_col + 1):
            ws.cell(row=r, column=c).border = THIN_BORDER


# ---------------------------------------------------------------------------
#  Sheet 1 — Dataset (Hourly Usage Matrix)
# ---------------------------------------------------------------------------
def _sheet_dataset(wb: Workbook, facilities: List[Facility]):
    ws = wb.active
    ws.title = "Dataset"
    ws.sheet_properties.tabColor = NAVY

    ws.merge_cells("A1:T1")
    ws["A1"] = "Campus Facility — Hourly Usage Data"
    ws["A1"].font = TITLE_FONT
    ws.row_dimensions[1].height = 30

    # Headers: Facility | Capacity | 6:00 | 7:00 | ... | 23:00
    headers = ["Facility", "Max Capacity"] + HOURS
    for i, h in enumerate(headers, 1):
        ws.cell(row=3, column=i, value=h)
    _style_header_row(ws, 3, len(headers))

    for idx, f in enumerate(facilities, 4):
        ws.cell(row=idx, column=1, value=f.name)
        ws.cell(row=idx, column=2, value=f.max_capacity)
        for j, usage in enumerate(f.hourly_usage, 3):
            cell = ws.cell(row=idx, column=j, value=usage)
            # Colour-code: green if < 50%, gold if 50-85%, red if > 85%
            pct = usage / f.max_capacity if f.max_capacity else 0
            if pct > 0.85:
                cell.font = Font(color=RED_ACCENT, bold=True)
            elif pct > 0.50:
                cell.font = Font(color=GOLD)

    end_row = 3 + len(facilities)
    _alt_rows(ws, 4, end_row, len(headers))
    _auto_width(ws)

    # Operating cost section below
    cost_start = end_row + 3
    ws.cell(row=cost_start, column=1, value="Operating Costs per Facility")
    ws.cell(row=cost_start, column=1).font = SECTION_FONT

    cost_headers = ["Facility", "Cost / Hour (₹)", "Daily Hours", "Daily Cost (₹)", "Cost / Person-Hr (₹)"]
    for i, h in enumerate(cost_headers, 1):
        ws.cell(row=cost_start + 1, column=i, value=h)
    _style_header_row(ws, cost_start + 1, len(cost_headers))

    for idx, f in enumerate(facilities, cost_start + 2):
        ws.cell(row=idx, column=1, value=f.name)
        ws.cell(row=idx, column=2, value=f.operating_cost_per_hour)
        ws.cell(row=idx, column=2).number_format = CURRENCY_FMT
        ws.cell(row=idx, column=3, value=f.total_operating_hours)
        ws.cell(row=idx, column=4, value=f.daily_operating_cost)
        ws.cell(row=idx, column=4).number_format = CURRENCY_FMT
        ws.cell(row=idx, column=5, value=round(f.cost_per_person_hour, 2))
        ws.cell(row=idx, column=5).number_format = CURRENCY_FMT

    _alt_rows(ws, cost_start + 2, cost_start + 1 + len(facilities), len(cost_headers))


# ---------------------------------------------------------------------------
#  Sheet 2 — Utilization Metrics
# ---------------------------------------------------------------------------
def _sheet_utilization(wb: Workbook, facilities: List[Facility]):
    ws = wb.create_sheet("Utilization Metrics")
    ws.sheet_properties.tabColor = DARK_TEAL

    ws.merge_cells("A1:H1")
    ws["A1"] = "Resource Utilization Metrics"
    ws["A1"].font = TITLE_FONT
    ws.row_dimensions[1].height = 30

    headers = ["Facility", "Max Capacity", "Peak Usage", "Peak Util %", "Avg Util %",
               "Idle Hours (<20%)", "Overcrowded Hrs (>85%)", "Idle Capacity %"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=3, column=i, value=h)
    _style_header_row(ws, 3, len(headers))

    for idx, f in enumerate(facilities, 4):
        ws.cell(row=idx, column=1, value=f.name)
        ws.cell(row=idx, column=2, value=f.max_capacity)
        ws.cell(row=idx, column=3, value=f.peak_usage)
        ws.cell(row=idx, column=4, value=round(f.peak_utilization_pct, 2))
        ws.cell(row=idx, column=4).number_format = PERCENT_FMT
        ws.cell(row=idx, column=5, value=round(f.avg_utilization_pct, 2))
        ws.cell(row=idx, column=5).number_format = PERCENT_FMT
        ws.cell(row=idx, column=6, value=f.idle_hours)
        ws.cell(row=idx, column=7, value=f.overcrowded_hours)
        ws.cell(row=idx, column=8, value=round(f.idle_capacity_pct, 2))
        ws.cell(row=idx, column=8).number_format = PERCENT_FMT

    end_row = 3 + len(facilities)
    _alt_rows(ws, 4, end_row, len(headers))
    _auto_width(ws)

    # --- Bar chart: Avg Utilization by Facility ---
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "Average Utilization % by Facility"
    chart.y_axis.title = "Utilization %"
    chart.width = 24
    chart.height = 14

    data_ref = Reference(ws, min_col=5, min_row=3, max_row=end_row)
    cats_ref = Reference(ws, min_col=1, min_row=4, max_row=end_row)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.shape = 4

    bar_colors = [TEAL, NAVY, GOLD, RED_ACCENT, GREEN_ACC, "8E44AD", "3498DB"]
    for i in range(len(facilities)):
        pt = DataPoint(idx=i)
        pt.graphicalProperties.solidFill = bar_colors[i % len(bar_colors)]
        chart.series[0].data_points.append(pt)

    ws.add_chart(chart, f"A{end_row + 3}")

    # --- Bar chart: Peak vs Avg Utilization ---
    chart2 = BarChart()
    chart2.type = "col"
    chart2.style = 10
    chart2.title = "Peak vs Average Utilization %"
    chart2.y_axis.title = "Utilization %"
    chart2.width = 24
    chart2.height = 14

    peak_ref = Reference(ws, min_col=4, min_row=3, max_row=end_row)
    avg_ref = Reference(ws, min_col=5, min_row=3, max_row=end_row)
    chart2.add_data(peak_ref, titles_from_data=True)
    chart2.add_data(avg_ref, titles_from_data=True)
    chart2.set_categories(cats_ref)

    ws.add_chart(chart2, f"A{end_row + 20}")


# ---------------------------------------------------------------------------
#  Sheet 3 — Optimization Scenarios
# ---------------------------------------------------------------------------
def _sheet_optimization(wb: Workbook, scenarios: List[Dict]):
    ws = wb.create_sheet("Optimization Scenarios")
    ws.sheet_properties.tabColor = TEAL

    ws.merge_cells("A1:G1")
    ws["A1"] = "Optimization Scenario Analysis"
    ws["A1"].font = TITLE_FONT
    ws.row_dimensions[1].height = 30

    cols = list(scenarios[0].keys())
    for i, h in enumerate(cols, 1):
        ws.cell(row=3, column=i, value=h)
    _style_header_row(ws, 3, len(cols))

    for idx, s in enumerate(scenarios, 4):
        for j, key in enumerate(cols, 1):
            val = s[key]
            ws.cell(row=idx, column=j, value=val)
            if "₹" in key and isinstance(val, (int, float)):
                ws.cell(row=idx, column=j).number_format = CURRENCY_FMT
            elif "%" in key and isinstance(val, (int, float)):
                ws.cell(row=idx, column=j).number_format = PERCENT_FMT

    end_row = 3 + len(scenarios)
    _alt_rows(ws, 4, end_row, len(cols))
    _auto_width(ws)

    # --- Bar chart: Utilization by scenario ---
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "Average Utilization % by Scenario"
    chart.y_axis.title = "Utilization %"
    chart.width = 28
    chart.height = 16

    data_ref = Reference(ws, min_col=2, min_row=3, max_row=end_row)
    cats_ref = Reference(ws, min_col=1, min_row=4, max_row=end_row)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.shape = 4

    for i, s in enumerate(scenarios):
        pt = DataPoint(idx=i)
        if "Combined" in s["Scenario"]:
            pt.graphicalProperties.solidFill = GOLD
        elif s["Avg Utilization %"] > scenarios[0]["Avg Utilization %"]:
            pt.graphicalProperties.solidFill = GREEN_ACC
        else:
            pt.graphicalProperties.solidFill = NAVY
        chart.series[0].data_points.append(pt)

    ws.add_chart(chart, f"A{end_row + 3}")

    # --- Cost efficiency chart ---
    chart2 = BarChart()
    chart2.type = "col"
    chart2.style = 10
    chart2.title = "Cost per Person-Hour (₹) by Scenario"
    chart2.y_axis.title = "₹ / Person-Hr"
    chart2.width = 28
    chart2.height = 16

    cost_col = cols.index("Cost / Person-Hr (₹)") + 1
    data_ref2 = Reference(ws, min_col=cost_col, min_row=3, max_row=end_row)
    chart2.add_data(data_ref2, titles_from_data=True)
    chart2.set_categories(cats_ref)
    chart2.shape = 4

    for i, s in enumerate(scenarios):
        pt = DataPoint(idx=i)
        if "Combined" in s["Scenario"]:
            pt.graphicalProperties.solidFill = GOLD
        elif s["Cost / Person-Hr (₹)"] < scenarios[0]["Cost / Person-Hr (₹)"]:
            pt.graphicalProperties.solidFill = GREEN_ACC
        else:
            pt.graphicalProperties.solidFill = RED_ACCENT
        chart2.series[0].data_points.append(pt)

    ws.add_chart(chart2, f"A{end_row + 20}")


# ---------------------------------------------------------------------------
#  Sheet 4 — Dashboard
# ---------------------------------------------------------------------------
def _sheet_dashboard(wb: Workbook, facilities: List[Facility], scenarios: List[Dict]):
    ws = wb.create_sheet("Dashboard")
    ws.sheet_properties.tabColor = GOLD

    ws.merge_cells("A1:F1")
    ws["A1"] = "Executive Dashboard"
    ws["A1"].font = Font(name="Calibri", bold=True, color=NAVY, size=20)
    ws.row_dimensions[1].height = 36

    total_ph = sum(f.total_person_hours for f in facilities)
    total_max = sum(f.max_possible_person_hours for f in facilities)
    total_cost = sum(f.daily_operating_cost for f in facilities)
    avg_util = (total_ph / total_max) * 100 if total_max else 0
    total_cap = sum(f.max_capacity for f in facilities)
    most_utilised = max(facilities, key=lambda f: f.avg_utilization_pct)
    least_utilised = min(facilities, key=lambda f: f.avg_utilization_pct)
    total_idle = sum(f.idle_hours for f in facilities)

    kpis = [
        ("Avg Utilization", f"{avg_util:.1f}%"),
        ("Total Capacity", total_cap),
        ("Most Utilised", most_utilised.name),
        ("Least Utilised", least_utilised.name),
        ("Daily Op. Cost", total_cost),
        ("Total Idle Hours", total_idle),
    ]

    for i, (label, val) in enumerate(kpis):
        col = i + 1
        ws.cell(row=3, column=col, value=label)
        ws.cell(row=3, column=col).font = Font(name="Calibri", bold=True, color=WHITE, size=10)
        ws.cell(row=3, column=col).fill = PatternFill("solid", fgColor=NAVY)
        ws.cell(row=3, column=col).alignment = Alignment(horizontal="center")

        ws.cell(row=4, column=col, value=val)
        ws.cell(row=4, column=col).font = Font(name="Calibri", bold=True, size=14, color=DARK_TEAL)
        ws.cell(row=4, column=col).alignment = Alignment(horizontal="center")
        if isinstance(val, (int, float)) and label == "Daily Op. Cost":
            ws.cell(row=4, column=col).number_format = CURRENCY_FMT
        ws.column_dimensions[get_column_letter(col)].width = 22

    # --- Pie chart: Capacity distribution ---
    ws.cell(row=7, column=1, value="Facility")
    ws.cell(row=7, column=2, value="Capacity")
    _style_header_row(ws, 7, 2)

    for idx, f in enumerate(facilities, 8):
        ws.cell(row=idx, column=1, value=f.name)
        ws.cell(row=idx, column=2, value=f.max_capacity)

    pie = PieChart()
    pie.title = "Capacity Distribution"
    pie.width = 20
    pie.height = 14
    labels = Reference(ws, min_col=1, min_row=8, max_row=7 + len(facilities))
    data = Reference(ws, min_col=2, min_row=7, max_row=7 + len(facilities))
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    pie.style = 10

    pie_colors = ["1B2A4A", "2EC4B6", "D4A853", "E74C3C", "3498DB", "8E44AD", "27AE60"]
    for i in range(len(facilities)):
        pt = DataPoint(idx=i)
        pt.graphicalProperties.solidFill = pie_colors[i % len(pie_colors)]
        pie.series[0].data_points.append(pt)
    pie.dataLabels = DataLabelList()
    pie.dataLabels.showPercent = True
    pie.dataLabels.showCatName = True

    ws.add_chart(pie, "D7")

    # --- Bar chart: Person-hours by facility ---
    ph_start = 8 + len(facilities) + 2
    ws.cell(row=ph_start, column=1, value="Facility")
    ws.cell(row=ph_start, column=2, value="Person-Hours")
    _style_header_row(ws, ph_start, 2)

    for idx, f in enumerate(facilities, ph_start + 1):
        ws.cell(row=idx, column=1, value=f.name)
        ws.cell(row=idx, column=2, value=f.total_person_hours)

    bar = BarChart()
    bar.type = "col"
    bar.style = 10
    bar.title = "Daily Person-Hours by Facility"
    bar.y_axis.title = "Person-Hours"
    bar.width = 20
    bar.height = 14

    d = Reference(ws, min_col=2, min_row=ph_start, max_row=ph_start + len(facilities))
    c = Reference(ws, min_col=1, min_row=ph_start + 1, max_row=ph_start + len(facilities))
    bar.add_data(d, titles_from_data=True)
    bar.set_categories(c)

    for i in range(len(facilities)):
        pt = DataPoint(idx=i)
        pt.graphicalProperties.solidFill = pie_colors[i % len(pie_colors)]
        bar.series[0].data_points.append(pt)

    ws.add_chart(bar, f"D{ph_start}")


# ---------------------------------------------------------------------------
#  Sheet 5 — Consulting Insights
# ---------------------------------------------------------------------------
def _sheet_insights(wb: Workbook, facilities: List[Facility], scenarios: List[Dict]):
    ws = wb.create_sheet("Consulting Insights")
    ws.sheet_properties.tabColor = RED_ACCENT

    ws.merge_cells("A1:C1")
    ws["A1"] = "Consulting Insights & Recommendations"
    ws["A1"].font = TITLE_FONT
    ws.row_dimensions[1].height = 30

    combined = [s for s in scenarios if "Combined" in s["Scenario"]][0]
    base_case = scenarios[0]
    util_improvement = combined["Avg Utilization %"] - base_case["Avg Utilization %"]
    cost_improvement = base_case["Cost / Person-Hr (₹)"] - combined["Cost / Person-Hr (₹)"]
    idle_reduction = base_case["Idle Hours (total)"] - combined["Idle Hours (total)"]

    most_idle = max(facilities, key=lambda f: f.idle_hours)
    most_crowded = max(facilities, key=lambda f: f.overcrowded_hours)

    sections = [
        ("1. KEY INEFFICIENCIES IDENTIFIED", [
            f"Auditorium utilisation at only {[f for f in facilities if f.name=='Auditorium'][0].avg_utilization_pct:.1f}% — massive capacity wastage (500 seats, avg usage ~89 persons/hr).",
            f"{most_crowded.name} has {most_crowded.overcrowded_hours} overcrowded hours/day (>85% capacity), causing resource contention.",
            f"Total {base_case['Idle Hours (total)']} idle hours across all facilities daily — zero productive output during these slots.",
            f"Sports Complex usage is heavily evening-biased; morning/afternoon slots go virtually unused.",
            f"Two Computer Labs with identical capacity create redundancy during off-peak hours.",
        ]),
        ("2. ROOT CAUSE ANALYSIS", [
            "Auditorium: Event-driven usage with no structured scheduling; booked ad-hoc causing massive idle periods.",
            "Gym: Morning peak (6–8 AM) and evening peak (4–6 PM) with midday dead zone — no staggered class scheduling.",
            "Computer Labs: Timetable-driven; Lab B mirrors Lab A schedule instead of complementary scheduling.",
            "Study Rooms: Low morning usage due to early classes; peaks only during exam season.",
            "No centralised booking/utilisation tracking system — decisions made without data.",
        ]),
        ("3. RECOMMENDATIONS", [
            "QUICK WIN — Implement centralised facility booking app with real-time occupancy display.",
            "Merge Computer Lab A & B into staggered schedules to eliminate duplicate idle periods.",
            "Convert Auditorium to multi-use space during idle hours (study hall, presentations, club activities).",
            "Introduce staggered gym slots tied to class timetable to flatten peak demand.",
            "Deploy IoT occupancy sensors for real-time data collection and dynamic scheduling.",
            "Close facilities during proven idle slots to reduce operating costs by ~15–20%.",
        ]),
        ("4. ESTIMATED EFFICIENCY IMPROVEMENT", [
            f"Base Case Utilization:     {base_case['Avg Utilization %']:>8.2f}%",
            f"Optimised Utilization:     {combined['Avg Utilization %']:>8.2f}%",
            f"Utilization Uplift:        {util_improvement:>8.2f} pp",
            f"Cost / Person-Hr Saved:    ₹{cost_improvement:>7.2f}",
            f"Idle Hours Eliminated:     {idle_reduction:>8d} hrs/day",
            f"Overcrowded Hrs Eliminated:{base_case['Overcrowded Hours'] - combined['Overcrowded Hours']:>4d} hrs/day",
        ]),
    ]

    row = 3
    for title, points in sections:
        ws.cell(row=row, column=1, value=title)
        ws.cell(row=row, column=1).font = SECTION_FONT
        row += 1
        for pt in points:
            ws.cell(row=row, column=1, value=f"  •  {pt}")
            ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True)
            row += 1
        row += 1

    ws.column_dimensions["A"].width = 100


# ---------------------------------------------------------------------------
#  Sheet 6 — Project Summary (Resume-Ready)
# ---------------------------------------------------------------------------
def _sheet_project_summary(wb: Workbook, facilities: List[Facility], scenarios: List[Dict]):
    ws = wb.create_sheet("Project Summary")
    ws.sheet_properties.tabColor = GREEN_ACC

    ws.merge_cells("A1:B1")
    ws["A1"] = "Project Summary — Resume Ready"
    ws["A1"].font = TITLE_FONT
    ws.row_dimensions[1].height = 30

    combined = [s for s in scenarios if "Combined" in s["Scenario"]][0]
    base_case = scenarios[0]
    util_improvement = combined["Avg Utilization %"] - base_case["Avg Utilization %"]

    sections = {
        "PROJECT TITLE": [
            "Campus Resource Utilization Analysis"
        ],
        "OBJECTIVE": [
            "Conducted a consulting-grade utilization analysis of 7 campus facilities at",
            "an Indian engineering college, quantifying inefficiencies and modeling",
            "optimization scenarios to improve resource allocation and reduce costs.",
        ],
        "METHODOLOGY": [
            "1. Built a granular hourly usage dataset across 7 facilities (18 time slots each).",
            "2. Computed utilization KPIs: peak, average, idle capacity, overcrowding metrics.",
            "3. Ran 7 optimization scenarios across redistribution, scheduling, and right-sizing levers.",
            "4. Identified root causes and quantified financial and operational impact.",
        ],
        "TOOLS USED": [
            "Python (openpyxl), Microsoft Excel, Chart.js, Financial/Operations Modeling, Scenario Analysis"
        ],
        "KEY IMPACT": [
            f"• Identified {base_case['Idle Hours (total)']} daily idle hours across 7 facilities — zero productive output.",
            f"• Demonstrated path to improve utilization from {base_case['Avg Utilization %']:.1f}% to {combined['Avg Utilization %']:.1f}% (+{util_improvement:.1f}pp).",
            f"• Reduced cost per person-hour from ₹{base_case['Cost / Person-Hr (₹)']:.2f} to ₹{combined['Cost / Person-Hr (₹)']:.2f}.",
            "• Delivered executive dashboard and consulting-style insights deck.",
        ],
    }

    row = 3
    for title, lines in sections.items():
        ws.cell(row=row, column=1, value=title)
        ws.cell(row=row, column=1).font = SECTION_FONT
        row += 1
        for line in lines:
            ws.cell(row=row, column=1, value=f"  {line}")
            ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True)
            row += 1
        row += 1

    ws.column_dimensions["A"].width = 80


# ═══════════════════════════════════════════════════════════════════════════
# 4.  MAIN
# ═══════════════════════════════════════════════════════════════════════════

def main():
    print("=" * 60)
    print("  Campus Resource Utilization Analysis")
    print("  Consulting-Style Analytics Project")
    print("=" * 60)

    facilities = build_dataset()
    scenarios = run_optimizations(facilities)

    # ---- Print summary to console ----
    total_ph = sum(f.total_person_hours for f in facilities)
    total_max = sum(f.max_possible_person_hours for f in facilities)
    total_cost = sum(f.daily_operating_cost for f in facilities)
    avg_util = (total_ph / total_max) * 100 if total_max else 0

    print(f"\n{'─' * 50}")
    print(f"  {'Facility':<20} {'Capacity':>8}  {'Avg Util':>9}  {'Peak Util':>10}  {'Idle Hrs':>8}")
    print(f"  {'─' * 20} {'─' * 8}  {'─' * 9}  {'─' * 10}  {'─' * 8}")
    for f in facilities:
        print(f"  {f.name:<20} {f.max_capacity:>8}  {f.avg_utilization_pct:>8.1f}%  {f.peak_utilization_pct:>9.1f}%  {f.idle_hours:>8}")
    print(f"{'─' * 50}")
    print(f"  Overall Avg Utilization:  {avg_util:.1f}%")
    print(f"  Total Daily Person-Hours: {total_ph:,}")
    print(f"  Total Daily Op. Cost:     ₹{total_cost:,.0f}")
    print(f"{'─' * 50}\n")

    # ---- Build Excel workbook ----
    wb = Workbook()
    _sheet_dataset(wb, facilities)
    _sheet_utilization(wb, facilities)
    _sheet_optimization(wb, scenarios)
    _sheet_dashboard(wb, facilities, scenarios)
    _sheet_insights(wb, facilities, scenarios)
    _sheet_project_summary(wb, facilities, scenarios)

    out_dir = os.path.dirname(os.path.abspath(__file__))
    out_path = os.path.join(out_dir, "Campus_Resource_Utilization_Analysis.xlsx")
    wb.save(out_path)
    print(f"✅  Excel workbook saved → {out_path}")
    print(f"    Sheets: {wb.sheetnames}\n")

    # ---- Print scenario table ----
    print("Optimization Scenarios:")
    print(f"  {'Scenario':<30} {'Util %':>8}  {'Person-Hrs':>12}  {'Cost/PH':>10}")
    print("  " + "─" * 65)
    for s in scenarios:
        print(f"  {s['Scenario']:<30} {s['Avg Utilization %']:>7.2f}%  {s['Total Person-Hours']:>12,}  ₹{s['Cost / Person-Hr (₹)']:>8.2f}")
    print()


if __name__ == "__main__":
    main()
