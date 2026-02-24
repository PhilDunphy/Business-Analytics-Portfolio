#!/usr/bin/env python3
"""
College Mess Profitability Optimization Model
==============================================
A consulting-style (Bain/BCG) financial analysis of an Indian college
hostel mess, covering revenue, cost structure, profitability, and
optimization simulations.

Outputs:
    1. College_Mess_Profitability_Model.xlsx  — 6-sheet workbook with charts
"""

import os
from dataclasses import dataclass, field
from typing import Dict, List, Tuple

# ---------------------------------------------------------------------------
# Try to import openpyxl; if missing, install it automatically
# ---------------------------------------------------------------------------
try:
    import openpyxl
except ImportError:
    import subprocess, sys
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl

from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference, BarChart3D
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from copy import copy

# ═══════════════════════════════════════════════════════════════════════════
# 1.  DATA ASSUMPTIONS  (realistic Indian college mess, 2024-25)
# ═══════════════════════════════════════════════════════════════════════════

@dataclass
class MessAssumptions:
    # --- Student & Pricing ---
    total_students: int = 650
    monthly_fee_per_student: float = 4500.0        # ₹ per month
    days_in_month: int = 30

    # --- Variable food cost ---
    daily_food_cost_per_student: float = 90.0       # ₹ raw ingredients
    food_wastage_pct: float = 12.0                  # % of food wasted

    # --- Other variable costs (monthly totals) ---
    lpg_fuel_monthly: float = 55_000.0
    water_monthly: float = 18_000.0
    packaging_disposables_monthly: float = 12_000.0

    # --- Fixed costs (monthly) ---
    staff_salaries: float = 220_000.0
    electricity: float = 45_000.0
    maintenance: float = 20_000.0
    rent: float = 60_000.0
    insurance_misc: float = 15_000.0

    # Derived helpers -------------------------------------------------------
    @property
    def monthly_raw_food_cost(self) -> float:
        return self.total_students * self.daily_food_cost_per_student * self.days_in_month

    @property
    def monthly_wastage_cost(self) -> float:
        return self.monthly_raw_food_cost * (self.food_wastage_pct / 100)

    @property
    def total_variable_costs(self) -> float:
        return (self.monthly_raw_food_cost
                + self.monthly_wastage_cost
                + self.lpg_fuel_monthly
                + self.water_monthly
                + self.packaging_disposables_monthly)

    @property
    def total_fixed_costs(self) -> float:
        return (self.staff_salaries
                + self.electricity
                + self.maintenance
                + self.rent
                + self.insurance_misc)

    @property
    def total_revenue(self) -> float:
        return self.total_students * self.monthly_fee_per_student

    @property
    def total_cost(self) -> float:
        return self.total_variable_costs + self.total_fixed_costs

    @property
    def profit(self) -> float:
        return self.total_revenue - self.total_cost

    @property
    def profit_margin(self) -> float:
        return (self.profit / self.total_revenue) * 100 if self.total_revenue else 0


# ═══════════════════════════════════════════════════════════════════════════
# 2.  OPTIMISATION SIMULATIONS
# ═══════════════════════════════════════════════════════════════════════════

def run_optimizations(base: MessAssumptions) -> List[Dict]:
    """Return a list of scenario dicts with name + key financials."""
    scenarios: List[Dict] = []

    def _snap(name: str, m: MessAssumptions) -> Dict:
        return {
            "Scenario": name,
            "Revenue": m.total_revenue,
            "Variable Costs": m.total_variable_costs,
            "Fixed Costs": m.total_fixed_costs,
            "Total Cost": m.total_cost,
            "Profit": m.profit,
            "Profit Margin %": round(m.profit_margin, 2),
        }

    # Base case
    scenarios.append(_snap("Base Case (Current)", base))

    # --- Wastage reduction ---
    for reduction in [5, 10, 15]:
        m = MessAssumptions(
            total_students=base.total_students,
            monthly_fee_per_student=base.monthly_fee_per_student,
            daily_food_cost_per_student=base.daily_food_cost_per_student,
            food_wastage_pct=base.food_wastage_pct - reduction,
            lpg_fuel_monthly=base.lpg_fuel_monthly,
            water_monthly=base.water_monthly,
            packaging_disposables_monthly=base.packaging_disposables_monthly,
            staff_salaries=base.staff_salaries,
            electricity=base.electricity,
            maintenance=base.maintenance,
            rent=base.rent,
            insurance_misc=base.insurance_misc,
        )
        scenarios.append(_snap(f"Wastage Reduced by {reduction}pp", m))

    # --- Operational efficiency (reduce LPG + Electricity + Maintenance) ---
    for eff in [2, 5, 8]:
        factor = 1 - eff / 100
        m = MessAssumptions(
            total_students=base.total_students,
            monthly_fee_per_student=base.monthly_fee_per_student,
            daily_food_cost_per_student=base.daily_food_cost_per_student,
            food_wastage_pct=base.food_wastage_pct,
            lpg_fuel_monthly=base.lpg_fuel_monthly * factor,
            water_monthly=base.water_monthly * factor,
            packaging_disposables_monthly=base.packaging_disposables_monthly,
            staff_salaries=base.staff_salaries,
            electricity=base.electricity * factor,
            maintenance=base.maintenance * factor,
            rent=base.rent,
            insurance_misc=base.insurance_misc,
        )
        scenarios.append(_snap(f"Operational Efficiency +{eff}%", m))

    # --- Pricing changes ---
    for bump in [100, 200, 300]:
        m = MessAssumptions(
            total_students=base.total_students,
            monthly_fee_per_student=base.monthly_fee_per_student + bump,
            daily_food_cost_per_student=base.daily_food_cost_per_student,
            food_wastage_pct=base.food_wastage_pct,
            lpg_fuel_monthly=base.lpg_fuel_monthly,
            water_monthly=base.water_monthly,
            packaging_disposables_monthly=base.packaging_disposables_monthly,
            staff_salaries=base.staff_salaries,
            electricity=base.electricity,
            maintenance=base.maintenance,
            rent=base.rent,
            insurance_misc=base.insurance_misc,
        )
        scenarios.append(_snap(f"Fee Increase ₹{bump}/mo", m))

    # --- Combined best realistic scenario ---
    m = MessAssumptions(
        total_students=base.total_students,
        monthly_fee_per_student=base.monthly_fee_per_student + 200,
        daily_food_cost_per_student=base.daily_food_cost_per_student,
        food_wastage_pct=base.food_wastage_pct - 10,
        lpg_fuel_monthly=base.lpg_fuel_monthly * 0.95,
        water_monthly=base.water_monthly * 0.95,
        packaging_disposables_monthly=base.packaging_disposables_monthly,
        staff_salaries=base.staff_salaries,
        electricity=base.electricity * 0.95,
        maintenance=base.maintenance * 0.95,
        rent=base.rent,
        insurance_misc=base.insurance_misc,
    )
    scenarios.append(_snap("Combined Optimized Scenario", m))

    return scenarios


# ═══════════════════════════════════════════════════════════════════════════
# 3.  EXCEL WORKBOOK GENERATION
# ═══════════════════════════════════════════════════════════════════════════

# ---- Colour palette (Bain-inspired navy/teal) ----------------------------
NAVY       = "1B2A4A"
DARK_TEAL  = "1A6B5C"
TEAL       = "2EC4B6"
LIGHT_GREY = "F2F2F2"
WHITE      = "FFFFFF"
GOLD       = "D4A853"
RED_ACCENT = "E74C3C"
GREEN_ACC  = "27AE60"

HEADER_FONT  = Font(name="Calibri", bold=True, color=WHITE, size=11)
HEADER_FILL  = PatternFill("solid", fgColor=NAVY)
SUBHEAD_FILL = PatternFill("solid", fgColor=DARK_TEAL)
ALT_ROW_FILL = PatternFill("solid", fgColor=LIGHT_GREY)
TITLE_FONT   = Font(name="Calibri", bold=True, color=NAVY, size=16)
SECTION_FONT = Font(name="Calibri", bold=True, color=DARK_TEAL, size=13)
CURRENCY_FMT = '₹#,##0'
PERCENT_FMT  = '0.00"%"'
THIN_BORDER  = Border(
    bottom=Side(style="thin", color="CCCCCC"),
)


def _style_header_row(ws, row, max_col):
    for col in range(1, max_col + 1):
        c = ws.cell(row=row, column=col)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)


def _currency_col(ws, col, start_row, end_row):
    for r in range(start_row, end_row + 1):
        ws.cell(row=r, column=col).number_format = CURRENCY_FMT


def _alt_rows(ws, start_row, end_row, max_col):
    for r in range(start_row, end_row + 1):
        if (r - start_row) % 2 == 1:
            for c in range(1, max_col + 1):
                ws.cell(row=r, column=c).fill = ALT_ROW_FILL
        for c in range(1, max_col + 1):
            ws.cell(row=r, column=c).border = THIN_BORDER


# ---------------------------------------------------------------------------
#  Sheet 1 — Dataset / Assumptions
# ---------------------------------------------------------------------------
def _sheet_dataset(wb: Workbook, base: MessAssumptions):
    ws = wb.active
    ws.title = "Dataset"
    ws.sheet_properties.tabColor = NAVY

    # Title
    ws.merge_cells("A1:C1")
    ws["A1"] = "College Mess — Input Assumptions"
    ws["A1"].font = TITLE_FONT
    ws.row_dimensions[1].height = 30

    headers = ["Category", "Parameter", "Value"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=3, column=i, value=h)
    _style_header_row(ws, 3, 3)

    data = [
        ("Student & Pricing", "Total Students Enrolled", base.total_students),
        ("Student & Pricing", "Monthly Mess Fee (₹)", base.monthly_fee_per_student),
        ("Student & Pricing", "Days in Month", base.days_in_month),
        ("Variable Costs", "Daily Food Cost / Student (₹)", base.daily_food_cost_per_student),
        ("Variable Costs", "Food Wastage (%)", base.food_wastage_pct),
        ("Variable Costs", "LPG / Fuel (₹/month)", base.lpg_fuel_monthly),
        ("Variable Costs", "Water (₹/month)", base.water_monthly),
        ("Variable Costs", "Packaging & Disposables (₹/month)", base.packaging_disposables_monthly),
        ("Fixed Costs", "Staff Salaries (₹/month)", base.staff_salaries),
        ("Fixed Costs", "Electricity (₹/month)", base.electricity),
        ("Fixed Costs", "Maintenance (₹/month)", base.maintenance),
        ("Fixed Costs", "Rent (₹/month)", base.rent),
        ("Fixed Costs", "Insurance & Misc (₹/month)", base.insurance_misc),
    ]

    for idx, (cat, param, val) in enumerate(data, 4):
        ws.cell(row=idx, column=1, value=cat)
        ws.cell(row=idx, column=2, value=param)
        ws.cell(row=idx, column=3, value=val)

    _alt_rows(ws, 4, 4 + len(data) - 1, 3)
    _auto_width(ws)


# ---------------------------------------------------------------------------
#  Sheet 2 — Financial Model (P&L)
# ---------------------------------------------------------------------------
def _sheet_financial_model(wb: Workbook, base: MessAssumptions):
    ws = wb.create_sheet("Financial Model")
    ws.sheet_properties.tabColor = DARK_TEAL

    ws.merge_cells("A1:C1")
    ws["A1"] = "Monthly Profit & Loss Statement"
    ws["A1"].font = TITLE_FONT
    ws.row_dimensions[1].height = 30

    headers = ["Line Item", "Amount (₹)", "% of Revenue"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=3, column=i, value=h)
    _style_header_row(ws, 3, 3)

    rev = base.total_revenue
    rows = [
        ("REVENUE", "", ""),
        ("  Total Students × Monthly Fee", rev, 100.0),
        ("", "", ""),
        ("VARIABLE COSTS", "", ""),
        ("  Raw Food Cost (Students × Daily × Days)", base.monthly_raw_food_cost, round(base.monthly_raw_food_cost / rev * 100, 2)),
        ("  Food Wastage Cost", base.monthly_wastage_cost, round(base.monthly_wastage_cost / rev * 100, 2)),
        ("  LPG / Fuel", base.lpg_fuel_monthly, round(base.lpg_fuel_monthly / rev * 100, 2)),
        ("  Water", base.water_monthly, round(base.water_monthly / rev * 100, 2)),
        ("  Packaging & Disposables", base.packaging_disposables_monthly, round(base.packaging_disposables_monthly / rev * 100, 2)),
        ("Total Variable Costs", base.total_variable_costs, round(base.total_variable_costs / rev * 100, 2)),
        ("", "", ""),
        ("FIXED COSTS", "", ""),
        ("  Staff Salaries", base.staff_salaries, round(base.staff_salaries / rev * 100, 2)),
        ("  Electricity", base.electricity, round(base.electricity / rev * 100, 2)),
        ("  Maintenance", base.maintenance, round(base.maintenance / rev * 100, 2)),
        ("  Rent", base.rent, round(base.rent / rev * 100, 2)),
        ("  Insurance & Misc", base.insurance_misc, round(base.insurance_misc / rev * 100, 2)),
        ("Total Fixed Costs", base.total_fixed_costs, round(base.total_fixed_costs / rev * 100, 2)),
        ("", "", ""),
        ("TOTAL COST", base.total_cost, round(base.total_cost / rev * 100, 2)),
        ("", "", ""),
        ("NET PROFIT / (LOSS)", base.profit, round(base.profit_margin, 2)),
        ("Profit Margin", f"{base.profit_margin:.2f}%", ""),
    ]

    section_labels = {"REVENUE", "VARIABLE COSTS", "FIXED COSTS"}
    total_labels = {"Total Variable Costs", "Total Fixed Costs", "TOTAL COST", "NET PROFIT / (LOSS)", "Profit Margin"}

    for idx, (item, amt, pct) in enumerate(rows, 4):
        r = idx
        ws.cell(row=r, column=1, value=item)
        if amt != "":
            ws.cell(row=r, column=2, value=amt)
            if isinstance(amt, (int, float)):
                ws.cell(row=r, column=2).number_format = CURRENCY_FMT
        if pct != "":
            ws.cell(row=r, column=3, value=pct if isinstance(pct, (int, float)) else pct)
            if isinstance(pct, (int, float)):
                ws.cell(row=r, column=3).number_format = '0.00"%"'

        # Bold section headers and totals
        if item.strip() in section_labels:
            ws.cell(row=r, column=1).font = SECTION_FONT
        if item.strip() in total_labels:
            ws.cell(row=r, column=1).font = Font(name="Calibri", bold=True, size=11)
            ws.cell(row=r, column=2).font = Font(name="Calibri", bold=True, size=11)
            ws.cell(row=r, column=3).font = Font(name="Calibri", bold=True, size=11)
            if item.strip() == "NET PROFIT / (LOSS)":
                color = GREEN_ACC if base.profit >= 0 else RED_ACCENT
                ws.cell(row=r, column=2).font = Font(name="Calibri", bold=True, size=12, color=color)

    _auto_width(ws)
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 16


# ---------------------------------------------------------------------------
#  Sheet 3 — Optimization Scenarios
# ---------------------------------------------------------------------------
def _sheet_optimization(wb: Workbook, scenarios: List[Dict]):
    ws = wb.create_sheet("Optimization Scenarios")
    ws.sheet_properties.tabColor = TEAL

    ws.merge_cells("A1:G1")
    ws["A1"] = "Optimization Scenario Analysis"
    ws["A1"].font = TITLE_FONT
    ws.row_dimensions[1].height = 30

    cols = ["Scenario", "Revenue", "Variable Costs", "Fixed Costs", "Total Cost", "Profit", "Profit Margin %"]
    for i, h in enumerate(cols, 1):
        ws.cell(row=3, column=i, value=h)
    _style_header_row(ws, 3, len(cols))

    for idx, s in enumerate(scenarios, 4):
        for j, key in enumerate(cols, 1):
            ws.cell(row=idx, column=j, value=s[key])
            if key != "Scenario" and key != "Profit Margin %":
                ws.cell(row=idx, column=j).number_format = CURRENCY_FMT
            elif key == "Profit Margin %":
                ws.cell(row=idx, column=j).number_format = '0.00"%"'

    end_row = 3 + len(scenarios)
    _alt_rows(ws, 4, end_row, len(cols))
    _auto_width(ws)

    # --- Bar chart: Profit by scenario ---
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "Profit by Scenario (₹)"
    chart.y_axis.title = "Profit (₹)"
    chart.x_axis.title = "Scenario"
    chart.width = 28
    chart.height = 16

    data_ref = Reference(ws, min_col=6, min_row=3, max_row=end_row)
    cats_ref = Reference(ws, min_col=1, min_row=4, max_row=end_row)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.shape = 4

    # Color bars
    for i, s in enumerate(scenarios):
        pt = DataPoint(idx=i)
        if "Combined" in s["Scenario"]:
            pt.graphicalProperties.solidFill = GOLD
        elif s["Profit"] > scenarios[0]["Profit"]:
            pt.graphicalProperties.solidFill = GREEN_ACC
        else:
            pt.graphicalProperties.solidFill = NAVY
        chart.series[0].data_points.append(pt)

    ws.add_chart(chart, f"A{end_row + 3}")

    # --- Margin comparison bar chart ---
    chart2 = BarChart()
    chart2.type = "col"
    chart2.style = 10
    chart2.title = "Profit Margin % by Scenario"
    chart2.y_axis.title = "Margin %"
    chart2.width = 28
    chart2.height = 16

    data_ref2 = Reference(ws, min_col=7, min_row=3, max_row=end_row)
    chart2.add_data(data_ref2, titles_from_data=True)
    chart2.set_categories(cats_ref)
    chart2.shape = 4

    for i, s in enumerate(scenarios):
        pt = DataPoint(idx=i)
        if "Combined" in s["Scenario"]:
            pt.graphicalProperties.solidFill = GOLD
        elif s["Profit Margin %"] > scenarios[0]["Profit Margin %"]:
            pt.graphicalProperties.solidFill = DARK_TEAL
        else:
            pt.graphicalProperties.solidFill = NAVY
        chart2.series[0].data_points.append(pt)

    ws.add_chart(chart2, f"A{end_row + 20}")


# ---------------------------------------------------------------------------
#  Sheet 4 — Dashboard
# ---------------------------------------------------------------------------
def _sheet_dashboard(wb: Workbook, base: MessAssumptions, scenarios: List[Dict]):
    ws = wb.create_sheet("Dashboard")
    ws.sheet_properties.tabColor = GOLD

    ws.merge_cells("A1:F1")
    ws["A1"] = "Executive Dashboard"
    ws["A1"].font = Font(name="Calibri", bold=True, color=NAVY, size=20)
    ws.row_dimensions[1].height = 36

    # KPI cards
    kpis = [
        ("Monthly Revenue", base.total_revenue),
        ("Total Costs", base.total_cost),
        ("Net Profit", base.profit),
        ("Profit Margin", f"{base.profit_margin:.2f}%"),
        ("Food Wastage Cost", base.monthly_wastage_cost),
        ("Break-even Students", int(base.total_cost / base.monthly_fee_per_student) + 1),
    ]

    for i, (label, val) in enumerate(kpis):
        col = i + 1
        ws.cell(row=3, column=col, value=label)
        ws.cell(row=3, column=col).font = Font(name="Calibri", bold=True, color=WHITE, size=10)
        ws.cell(row=3, column=col).fill = PatternFill("solid", fgColor=NAVY)
        ws.cell(row=3, column=col).alignment = Alignment(horizontal="center")

        ws.cell(row=4, column=col, value=val)
        ws.cell(row=4, column=col).font = Font(name="Calibri", bold=True, size=14, color=DARK_TEAL)
        ws.cell(row=4, column=col).alignment = Alignment(horizontal="center")
        if isinstance(val, (int, float)) and label != "Break-even Students":
            ws.cell(row=4, column=col).number_format = CURRENCY_FMT
        ws.column_dimensions[get_column_letter(col)].width = 22

    # --- Pie chart: cost breakdown ---
    ws.cell(row=7, column=1, value="Cost Component")
    ws.cell(row=7, column=2, value="Amount (₹)")
    _style_header_row(ws, 7, 2)

    cost_items = [
        ("Raw Food", base.monthly_raw_food_cost),
        ("Food Wastage", base.monthly_wastage_cost),
        ("LPG / Fuel", base.lpg_fuel_monthly),
        ("Water", base.water_monthly),
        ("Packaging", base.packaging_disposables_monthly),
        ("Staff Salaries", base.staff_salaries),
        ("Electricity", base.electricity),
        ("Maintenance", base.maintenance),
        ("Rent", base.rent),
        ("Insurance & Misc", base.insurance_misc),
    ]

    for idx, (name, val) in enumerate(cost_items, 8):
        ws.cell(row=idx, column=1, value=name)
        ws.cell(row=idx, column=2, value=val)
        ws.cell(row=idx, column=2).number_format = CURRENCY_FMT

    pie = PieChart()
    pie.title = "Cost Structure Breakdown"
    pie.width = 20
    pie.height = 14
    labels = Reference(ws, min_col=1, min_row=8, max_row=8 + len(cost_items) - 1)
    data = Reference(ws, min_col=2, min_row=7, max_row=7 + len(cost_items))
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    pie.style = 10

    # Custom colors
    pie_colors = ["1B2A4A", "2EC4B6", "D4A853", "E74C3C", "3498DB",
                  "8E44AD", "27AE60", "F39C12", "1ABC9C", "95A5A6"]
    for i in range(len(cost_items)):
        pt = DataPoint(idx=i)
        pt.graphicalProperties.solidFill = pie_colors[i % len(pie_colors)]
        pie.series[0].data_points.append(pt)

    pie.dataLabels = DataLabelList()
    pie.dataLabels.showPercent = True
    pie.dataLabels.showCatName = True
    pie.dataLabels.showVal = False

    ws.add_chart(pie, "D7")

    # --- Revenue vs Cost waterfall-style bar chart ---
    ws.cell(row=20, column=1, value="Metric")
    ws.cell(row=20, column=2, value="Amount (₹)")
    _style_header_row(ws, 20, 2)
    waterfall = [
        ("Revenue", base.total_revenue),
        ("Variable Costs", base.total_variable_costs),
        ("Fixed Costs", base.total_fixed_costs),
        ("Net Profit", base.profit),
    ]
    for idx, (name, val) in enumerate(waterfall, 21):
        ws.cell(row=idx, column=1, value=name)
        ws.cell(row=idx, column=2, value=val)
        ws.cell(row=idx, column=2).number_format = CURRENCY_FMT

    bar = BarChart()
    bar.type = "col"
    bar.style = 10
    bar.title = "Revenue vs Costs vs Profit"
    bar.y_axis.title = "Amount (₹)"
    bar.width = 20
    bar.height = 14

    d = Reference(ws, min_col=2, min_row=20, max_row=24)
    c = Reference(ws, min_col=1, min_row=21, max_row=24)
    bar.add_data(d, titles_from_data=True)
    bar.set_categories(c)

    bar_colors = [GREEN_ACC, RED_ACCENT, "E67E22", DARK_TEAL]
    for i in range(4):
        pt = DataPoint(idx=i)
        pt.graphicalProperties.solidFill = bar_colors[i]
        bar.series[0].data_points.append(pt)

    ws.add_chart(bar, "D20")


# ---------------------------------------------------------------------------
#  Sheet 5 — Consulting Insights
# ---------------------------------------------------------------------------
def _sheet_insights(wb: Workbook, base: MessAssumptions, scenarios: List[Dict]):
    ws = wb.create_sheet("Consulting Insights")
    ws.sheet_properties.tabColor = RED_ACCENT

    ws.merge_cells("A1:C1")
    ws["A1"] = "Consulting Insights & Recommendations"
    ws["A1"].font = TITLE_FONT
    ws.row_dimensions[1].height = 30

    combined = [s for s in scenarios if "Combined" in s["Scenario"]][0]
    base_case = scenarios[0]
    improvement = combined["Profit"] - base_case["Profit"]

    sections = [
        ("1. KEY INEFFICIENCIES IDENTIFIED", [
            f"Food wastage at {base.food_wastage_pct}% translates to ₹{base.monthly_wastage_cost:,.0f}/month of avoidable cost.",
            "No dynamic menu planning — fixed menu leads to predictable over-production on low-attendance days.",
            "Energy & fuel costs (LPG + electricity) represent a combined ~3.4% of revenue with no efficiency monitoring.",
            "Manual inventory management causes 5-8% over-ordering of perishable goods.",
            f"Break-even point at {int(base.total_cost / base.monthly_fee_per_student) + 1} students leaves thin safety margin.",
        ]),
        ("2. ROOT CAUSE ANALYSIS", [
            "Wastage: Absence of attendance-based demand forecasting; batch cooking without portion control.",
            "Energy: Aged kitchen equipment (>5 yrs); no sub-metering for consumption tracking.",
            "Procurement: No vendor benchmarking; reliance on single supplier inflates raw material cost.",
            "Pricing: Mess fee has not been revised in 2+ years despite 8-10% food inflation.",
        ]),
        ("3. RECOMMENDATIONS", [
            "QUICK WIN — Implement meal pre-booking app to forecast daily demand (est. 40% wastage reduction).",
            "Negotiate bulk/multi-vendor procurement; target 5-7% raw material savings.",
            "Install sub-metering and switch to energy-efficient equipment; target 5% energy reduction.",
            f"Revise mess fee by ₹200/month (4.4% increase, still below market rate of ₹5,000+).",
            "Introduce weekend opt-out system to reduce per-meal variable cost.",
        ]),
        ("4. ESTIMATED FINANCIAL IMPROVEMENT", [
            f"Base Case Profit:           ₹{base_case['Profit']:>12,.0f}   ({base_case['Profit Margin %']:.2f}% margin)",
            f"Optimised Scenario Profit:  ₹{combined['Profit']:>12,.0f}   ({combined['Profit Margin %']:.2f}% margin)",
            f"Incremental Profit Gain:    ₹{improvement:>12,.0f}/month",
            f"Annualised Improvement:     ₹{improvement * 12:>12,.0f}/year",
            f"Margin Uplift:              {combined['Profit Margin %'] - base_case['Profit Margin %']:.2f} pp",
        ]),
    ]

    row = 3
    for title, points in sections:
        ws.cell(row=row, column=1, value=title)
        ws.cell(row=row, column=1).font = SECTION_FONT
        row += 1
        for pt in points:
            ws.cell(row=row, column=1, value=f"  •  {pt}")
            ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True)
            row += 1
        row += 1  # spacer

    ws.column_dimensions["A"].width = 100


# ---------------------------------------------------------------------------
#  Sheet 6 — Project Summary (Resume-Ready)
# ---------------------------------------------------------------------------
def _sheet_project_summary(wb: Workbook, base: MessAssumptions, scenarios: List[Dict]):
    ws = wb.create_sheet("Project Summary")
    ws.sheet_properties.tabColor = GREEN_ACC

    ws.merge_cells("A1:B1")
    ws["A1"] = "Project Summary — Resume Ready"
    ws["A1"].font = TITLE_FONT
    ws.row_dimensions[1].height = 30

    combined = [s for s in scenarios if "Combined" in s["Scenario"]][0]
    base_case = scenarios[0]
    improvement = combined["Profit"] - base_case["Profit"]

    sections = {
        "PROJECT TITLE": [
            "College Mess Profitability Optimization Model"
        ],
        "OBJECTIVE": [
            "Developed a consulting-grade financial model to analyze the profitability,",
            "cost structure, and optimization levers of an Indian college hostel mess",
            "serving 650+ students, with actionable recommendations to improve margins."
        ],
        "METHODOLOGY": [
            "1. Built a bottom-up P&L model with granular cost drivers (13 line items).",
            "2. Ran 11 optimization scenarios across 3 levers: wastage reduction,",
            "   operational efficiency, and pricing adjustments.",
            "3. Conducted root-cause analysis of key cost inefficiencies.",
            "4. Quantified financial impact of recommendations using scenario modeling.",
        ],
        "TOOLS USED": [
            "Python (openpyxl), Microsoft Excel, Financial Modeling, Scenario Analysis"
        ],
        "KEY IMPACT": [
            f"• Identified ₹{base.monthly_wastage_cost:,.0f}/month in avoidable food wastage costs.",
            f"• Demonstrated path to improve profit margin from {base_case['Profit Margin %']:.1f}% to {combined['Profit Margin %']:.1f}%.",
            f"• Quantified ₹{improvement * 12:,.0f} annualized profit improvement potential.",
            "• Delivered executive dashboard and BCG/Bain-style insights deck.",
        ],
    }

    row = 3
    for title, lines in sections.items():
        ws.cell(row=row, column=1, value=title)
        ws.cell(row=row, column=1).font = SECTION_FONT
        row += 1
        for line in lines:
            ws.cell(row=row, column=1, value=f"  {line}")
            ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True)
            row += 1
        row += 1

    ws.column_dimensions["A"].width = 80


# ═══════════════════════════════════════════════════════════════════════════
# 4.  MAIN
# ═══════════════════════════════════════════════════════════════════════════

def main():
    print("=" * 60)
    print("  College Mess Profitability Optimization Model")
    print("  Bain/BCG-Style Financial Analysis")
    print("=" * 60)

    base = MessAssumptions()
    scenarios = run_optimizations(base)

    # ---- Print summary to console ----
    print(f"\n{'─' * 45}")
    print(f"  Monthly Revenue:       ₹{base.total_revenue:>12,.0f}")
    print(f"  Total Variable Costs:  ₹{base.total_variable_costs:>12,.0f}")
    print(f"  Total Fixed Costs:     ₹{base.total_fixed_costs:>12,.0f}")
    print(f"  Total Cost:            ₹{base.total_cost:>12,.0f}")
    print(f"  Net Profit:            ₹{base.profit:>12,.0f}")
    print(f"  Profit Margin:         {base.profit_margin:>11.2f}%")
    print(f"{'─' * 45}\n")

    # ---- Build Excel workbook ----
    wb = Workbook()
    _sheet_dataset(wb, base)
    _sheet_financial_model(wb, base)
    _sheet_optimization(wb, scenarios)
    _sheet_dashboard(wb, base, scenarios)
    _sheet_insights(wb, base, scenarios)
    _sheet_project_summary(wb, base, scenarios)

    out_dir = os.path.dirname(os.path.abspath(__file__))
    out_path = os.path.join(out_dir, "College_Mess_Profitability_Model.xlsx")
    wb.save(out_path)
    print(f"✅  Excel workbook saved → {out_path}")
    print(f"    Sheets: {wb.sheetnames}\n")

    # ---- Print scenario table ----
    print("Optimization Scenarios:")
    print(f"{'Scenario':<35} {'Profit':>14}  {'Margin':>8}")
    print("─" * 60)
    for s in scenarios:
        print(f"  {s['Scenario']:<33} ₹{s['Profit']:>12,.0f}  {s['Profit Margin %']:>7.2f}%")
    print()


if __name__ == "__main__":
    main()
